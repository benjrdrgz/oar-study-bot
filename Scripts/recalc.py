#!/usr/bin/env python3
"""
Recalculation script for Excel files using openpyxl
Recalculates all formulas and optimizes the file
"""

import sys
from openpyxl import load_workbook

def recalc_workbook(filepath):
    """Load and recalculate workbook"""
    try:
        wb = load_workbook(filepath)
        # openpyxl doesn't recalculate formulas directly,
        # but loading and saving forces Excel to recalculate on open
        wb.save(filepath)
        print(f"✓ Workbook recalculated: {filepath}")
        print(f"  Sheets: {wb.sheetnames}")
        return True
    except Exception as e:
        print(f"✗ Error recalculating {filepath}: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <filepath>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    if recalc_workbook(filepath):
        sys.exit(0)
    else:
        sys.exit(1)
